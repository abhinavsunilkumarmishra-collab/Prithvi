#!/usr/bin/env python3
"""
Sync edits made in owner-land-loan-backend.xlsx back into data/plots-data.json,
which the PRTHVI Next.js app reads at build time.

Usage:
    python3 scripts/sync_excel_to_json.py

Run this from the project root after saving changes in the Excel file.
Geometry (map coordinates) is never touched by this script — only owner,
land, loan, and dispute fields are updated, matched by Plot ID.
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "owner-land-loan-backend.xlsx"
JSON_PATH = ROOT / "data" / "plots-data.json"


def sheet_rows(ws):
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        yield dict(zip(headers, row))


def main():
    if not EXCEL_PATH.exists():
        sys.exit(f"Excel file not found: {EXCEL_PATH}")
    if not JSON_PATH.exists():
        sys.exit(f"Data file not found: {JSON_PATH}")

    with open(JSON_PATH) as f:
        plots = json.load(f)
    by_id = {p["id"]: p for p in plots}

    wb = load_workbook(EXCEL_PATH, data_only=True)
    changed_ids = set()

    # Owners
    for row in sheet_rows(wb["Owners"]):
        p = by_id.get(row["Plot ID"])
        if not p:
            continue
        co = p["currentOwner"]
        new_name = row.get("Current Owner Name") or co["name"]
        new_phone = row.get("Owner Phone") or p["ownerPhone"]
        new_from = row.get("Owner Since") or co["from"]
        if (new_name, new_phone, str(new_from)) != (
            co["name"], p["ownerPhone"], str(co["from"])
        ):
            changed_ids.add(p["id"])
        co["name"] = str(new_name)
        p["ownerPhone"] = str(new_phone)
        co["from"] = str(new_from)

        prev_name = row.get("Previous Owner Name")
        if prev_name:
            prev = p["previousOwners"][0] if p["previousOwners"] else {}
            prev["name"] = str(prev_name)
            if row.get("Previous Owner From"):
                prev["from"] = str(row["Previous Owner From"])
            if row.get("Previous Owner To"):
                prev["to"] = str(row["Previous Owner To"])
            p["previousOwners"] = [prev]
        else:
            p["previousOwners"] = []

    # Land Info
    for row in sheet_rows(wb["Land Info"]):
        p = by_id.get(row["Plot ID"])
        if not p:
            continue
        p["surveyNumber"] = str(row.get("Survey Number") or p["surveyNumber"])
        p["gatNumber"] = str(row.get("Gat Number") or p["gatNumber"])
        p["village"] = str(row.get("Village") or p["village"])
        p["ward"] = str(row.get("Ward") or p["ward"])
        p["taluka"] = str(row.get("Taluka") or p["taluka"])
        p["district"] = str(row.get("District") or p["district"])
        p["state"] = str(row.get("State") or p["state"])
        if row.get("Land Type"):
            p["landType"] = str(row["Land Type"])
        if row.get("Area (sqm)") is not None:
            p["areaSqM"] = int(row["Area (sqm)"])
        if row.get("Verification Status"):
            p["verificationStatus"] = str(row["Verification Status"])
        if row.get("Last Verified"):
            p["lastVerified"] = str(row["Last Verified"])
        if row.get("Source Note"):
            p["sourceNote"] = str(row["Source Note"])

    # Loans
    for row in sheet_rows(wb["Loans"]):
        p = by_id.get(row["Plot ID"])
        if not p:
            continue
        if row.get("Lender"):
            p["loan"] = {
                "lender": str(row["Lender"]),
                "accountRef": str(row.get("Account Ref") or ""),
                "amountInr": int(row["Loan Amount (INR)"]) if row.get("Loan Amount (INR)") else 0,
                "sanctionedDate": str(row.get("Sanctioned Date") or ""),
                "status": str(row.get("Loan Status") or "active"),
            }
        else:
            p["loan"] = None

    # Disputes
    for row in sheet_rows(wb["Disputes"]):
        p = by_id.get(row["Plot ID"])
        if not p:
            continue
        if row.get("Case Number"):
            p["disputes"] = [{
                "caseNumber": str(row["Case Number"]),
                "court": str(row.get("Court") or ""),
                "filedDate": str(row.get("Filed Date") or ""),
                "status": str(row.get("Dispute Status") or "open"),
                "description": str(row.get("Description") or ""),
            }]
        else:
            p["disputes"] = []

    with open(JSON_PATH, "w") as f:
        json.dump(list(by_id.values()), f, indent=2)

    print(f"Synced {len(plots)} plots from Excel into {JSON_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
