import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("/home/claude/land-registry/project/data/plots-data.json") as f:
    plots = json.load(f)

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5233")
BODY_FONT = Font(name=FONT_NAME)
EDITABLE_FILL = PatternFill("solid", fgColor="FFF9C4")  # light yellow = safe to edit
LOCKED_FONT = Font(name=FONT_NAME, color="777777")  # id columns = don't touch

wb = Workbook()

def style_header(ws, headers, editable_cols):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 32

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------- Instructions sheet ----------------
ws = wb.active
ws.title = "Instructions"
lines = [
    ("PRTHVI Land Registry — Editable Backend Workbook", True, 14),
    ("", False, 11),
    ("This file is the editable backend for the land registry map/app.", False, 11),
    ("Each plot has a Plot ID (e.g. NGP-DHR-001) that links its rows across", False, 11),
    ("all sheets and links back to the plot shown on the map.", False, 11),
    ("", False, 11),
    ("HOW TO USE", True, 12),
    ("1. Open the sheet you want to change: Owners, Land Info, Loans, or Disputes.", False, 11),
    ("2. Find the row for the Plot ID you want to update (or use Ctrl+F to search a name).", False, 11),
    ("3. Edit only the yellow-highlighted columns. Never edit the 'Plot ID' column —", False, 11),
    ("   it is how this workbook stays linked to the map.", False, 11),
    ("4. Save the file.", False, 11),
    ("5. Send the saved file back so it can be re-synced into the app's data file", False, 11),
    ("   (data/plots-data.json). The map, plot pages, and search will then show", False, 11),
    ("   your changes.", False, 11),
    ("", False, 11),
    ("Example: to change who owns plot NGP-DHR-014, go to the Owners sheet, find", False, 11),
    ("row NGP-DHR-014, delete the name in 'Current Owner Name' and type the new name.", False, 11),
    ("", False, 11),
    ("SHEETS", True, 12),
    ("Owners      — current & previous owner name, relation, phone, since-date", False, 11),
    ("Land Info   — survey number, village, land type, area, verification status", False, 11),
    ("Loans       — lender, account, loan amount, status (blank = no loan on record)", False, 11),
    ("Disputes    — court case details, if any (blank = no dispute on record)", False, 11),
    ("", False, 11),
    ("Note: all names, phone numbers, loan accounts and case numbers in this demo", False, 10),
    ("dataset are synthetic placeholders, not real records.", False, 10),
]
for r, (text, bold, size) in enumerate(lines, start=1):
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size)
autosize(ws, [95])

# ---------------- Owners sheet ----------------
ws = wb.create_sheet("Owners")
headers = [
    "Plot ID", "Current Owner Name", "Owner Phone",
    "Owner Since", "Previous Owner Name",
    "Previous Owner From", "Previous Owner To",
]
style_header(ws, headers, editable_cols=range(2, 8))
for r, p in enumerate(plots, start=2):
    co = p["currentOwner"]
    prev = p["previousOwners"][0] if p["previousOwners"] else None
    row = [
        p["id"], co["name"], p["ownerPhone"], co["from"],
        prev["name"] if prev else "",
        prev["from"] if prev else "", prev.get("to", "") if prev else "",
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        if c == 1:
            cell.font = LOCKED_FONT
        else:
            cell.font = BODY_FONT
            cell.fill = EDITABLE_FILL
autosize(ws, [14, 22, 16, 13, 22, 15, 15])

# ---------------- Land Info sheet ----------------
ws = wb.create_sheet("Land Info")
headers = [
    "Plot ID", "Survey Number", "Gat Number", "Village", "Ward", "Taluka",
    "District", "State", "Land Type", "Area (sqm)", "Verification Status",
    "Last Verified", "Source Note",
]
style_header(ws, headers, editable_cols=range(2, 14))
for r, p in enumerate(plots, start=2):
    row = [
        p["id"], p["surveyNumber"], p["gatNumber"], p["village"], p["ward"],
        p["taluka"], p["district"], p["state"], p["landType"], p["areaSqM"],
        p["verificationStatus"], p["lastVerified"], p["sourceNote"],
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        if c == 1:
            cell.font = LOCKED_FONT
        else:
            cell.font = BODY_FONT
            cell.fill = EDITABLE_FILL
autosize(ws, [14, 16, 10, 13, 16, 13, 10, 12, 13, 12, 18, 13, 34])

# ---------------- Loans sheet ----------------
ws = wb.create_sheet("Loans")
headers = ["Plot ID", "Lender", "Account Ref", "Loan Amount (INR)", "Sanctioned Date", "Loan Status"]
style_header(ws, headers, editable_cols=range(2, 7))
for r, p in enumerate(plots, start=2):
    loan = p.get("loan")
    row = [
        p["id"],
        loan["lender"] if loan else "",
        loan["accountRef"] if loan else "",
        loan["amountInr"] if loan else "",
        loan["sanctionedDate"] if loan else "",
        loan["status"] if loan else "",
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        if c == 1:
            cell.font = LOCKED_FONT
        else:
            cell.font = BODY_FONT
            cell.fill = EDITABLE_FILL
autosize(ws, [14, 26, 22, 18, 16, 13])

# ---------------- Disputes sheet ----------------
ws = wb.create_sheet("Disputes")
headers = ["Plot ID", "Case Number", "Court", "Filed Date", "Dispute Status", "Description"]
style_header(ws, headers, editable_cols=range(2, 7))
for r, p in enumerate(plots, start=2):
    d = p["disputes"][0] if p.get("disputes") else None
    row = [
        p["id"],
        d["caseNumber"] if d else "",
        d["court"] if d else "",
        d["filedDate"] if d else "",
        d["status"] if d else "",
        d["description"] if d else "",
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        if c == 1:
            cell.font = LOCKED_FONT
        else:
            cell.font = BODY_FONT
            cell.fill = EDITABLE_FILL
autosize(ws, [14, 18, 26, 13, 15, 45])

out_path = "/home/claude/land-registry/project/owner-land-loan-backend.xlsx"
wb.save(out_path)
print("Saved", out_path, "with", len(plots), "plot rows per sheet")
